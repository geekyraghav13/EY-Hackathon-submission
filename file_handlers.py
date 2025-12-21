"""
File Handlers for Provider Validation System
Handles CSV/Excel upload parsing and report export (PDF/Excel)
"""
import csv
import json
import io
from datetime import datetime
from typing import Dict, List, Any, Optional


class FileUploadHandler:
    """Handles parsing of uploaded provider data files (CSV/Excel)"""
    
    REQUIRED_FIELDS = ['first_name', 'last_name', 'npi']
    OPTIONAL_FIELDS = [
        'provider_id', 'specialty', 'sub_specialty', 'phone', 'email',
        'address', 'city', 'state', 'zip_code', 'license_number',
        'license_status', 'license_expiry', 'board_certified',
        'years_experience', 'medical_school', 'accepting_new_patients',
        'languages', 'hospital_affiliations', 'insurance_accepted'
    ]
    
    @classmethod
    def parse_csv(cls, file_content: str) -> List[Dict[str, Any]]:
        """Parse CSV content into provider records"""
        providers = []
        reader = csv.DictReader(io.StringIO(file_content))
        
        for idx, row in enumerate(reader, start=1):
            provider = cls._normalize_row(row, idx)
            if provider:
                providers.append(provider)
        
        return providers
    
    @classmethod
    def parse_excel(cls, file_bytes: bytes) -> List[Dict[str, Any]]:
        """Parse Excel file into provider records"""
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(io.BytesIO(file_bytes))
            sheet = workbook.active
            
            headers = [cell.value.lower().replace(' ', '_') if cell.value else '' 
                      for cell in sheet[1]]
            
            providers = []
            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                row_dict = {headers[i]: val for i, val in enumerate(row) if i < len(headers)}
                provider = cls._normalize_row(row_dict, idx)
                if provider:
                    providers.append(provider)
            
            return providers
        except ImportError:
            raise ValueError("Excel support requires openpyxl. Install with: pip install openpyxl")
    
    @classmethod
    def _normalize_row(cls, row: Dict, idx: int) -> Optional[Dict[str, Any]]:
        """Normalize a row into a standard provider format"""
        # Normalize field names (lowercase, replace spaces with underscores)
        normalized = {}
        for key, value in row.items():
            if key:
                norm_key = key.lower().replace(' ', '_').strip()
                normalized[norm_key] = value
        
        # Check required fields
        for field in cls.REQUIRED_FIELDS:
            if field not in normalized or not normalized[field]:
                return None
        
        # Build provider record
        provider = {
            'provider_id': normalized.get('provider_id', f'PRV{idx:05d}'),
            'npi': str(normalized.get('npi', '')).strip(),
            'first_name': str(normalized.get('first_name', '')).strip(),
            'last_name': str(normalized.get('last_name', '')).strip(),
            'specialty': normalized.get('specialty', 'General Practice'),
            'sub_specialty': normalized.get('sub_specialty', 'General'),
            'phone': str(normalized.get('phone', '')).strip(),
            'email': str(normalized.get('email', '')).strip(),
            'address': str(normalized.get('address', '')).strip(),
            'city': str(normalized.get('city', '')).strip(),
            'state': str(normalized.get('state', '')).strip().upper()[:2],
            'zip_code': str(normalized.get('zip_code', '')).strip()[:5],
            'license_number': str(normalized.get('license_number', '')).strip(),
            'license_status': normalized.get('license_status', 'Unknown'),
            'license_expiry': normalized.get('license_expiry', ''),
            'board_certified': cls._parse_bool(normalized.get('board_certified', False)),
            'years_experience': cls._parse_int(normalized.get('years_experience', 0)),
            'medical_school': str(normalized.get('medical_school', '')).strip(),
            'accepting_new_patients': cls._parse_bool(normalized.get('accepting_new_patients', True)),
            'languages': cls._parse_list(normalized.get('languages', 'English')),
            'hospital_affiliations': cls._parse_list(normalized.get('hospital_affiliations', '')),
            'insurance_accepted': cls._parse_list(normalized.get('insurance_accepted', '')),
            'data_quality_score': 0,
            'last_verified': datetime.now().strftime('%Y-%m-%d'),
            'validation_status': 'Pending',
            'confidence_score': 0.0,
            'needs_manual_review': False,
            'issues_found': []
        }
        
        return provider
    
    @staticmethod
    def _parse_bool(value) -> bool:
        """Parse various boolean representations"""
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ('true', 'yes', '1', 'y')
        return bool(value)
    
    @staticmethod
    def _parse_int(value) -> int:
        """Parse integer with fallback to 0"""
        try:
            return int(value)
        except (ValueError, TypeError):
            return 0
    
    @staticmethod
    def _parse_list(value) -> List[str]:
        """Parse comma-separated string into list"""
        if isinstance(value, list):
            return value
        if isinstance(value, str) and value:
            return [item.strip() for item in value.split(',') if item.strip()]
        return []


class ReportExporter:
    """Handles exporting validation results to various formats"""
    
    @classmethod
    def to_excel(cls, results: Dict[str, Any]) -> bytes:
        """Export validation results to Excel format"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ValueError("Excel export requires openpyxl. Install with: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        cls._create_summary_sheet(ws_summary, results)
        
        # Provider Results Sheet
        ws_providers = wb.create_sheet("Provider Results")
        cls._create_providers_sheet(ws_providers, results)
        
        # Issues Sheet
        ws_issues = wb.create_sheet("Issues Found")
        cls._create_issues_sheet(ws_issues, results)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @classmethod
    def _create_summary_sheet(cls, ws, results: Dict):
        """Create summary worksheet"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Title
        ws['A1'] = "Provider Validation Summary Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Timestamp
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True)
        
        # Stats
        stats = results.get('processing_stats', {})
        summary = results.get('summary', {})
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        row = 4
        metrics = [
            ("Metric", "Value"),
            ("Total Providers", stats.get('total_providers', 0)),
            ("Providers Validated", stats.get('providers_validated', 0)),
            ("Providers Needing Review", stats.get('providers_needing_review', 0)),
            ("Average Quality Score", f"{summary.get('average_quality_score', 0):.1f}"),
            ("Processing Time", f"{stats.get('total_time_seconds', 0):.2f}s" if 'total_time_seconds' in stats else 'N/A'),
        ]
        
        for i, (metric, value) in enumerate(metrics):
            ws.cell(row=row+i, column=1, value=metric)
            ws.cell(row=row+i, column=2, value=value)
            if i == 0:
                ws.cell(row=row+i, column=1).fill = header_fill
                ws.cell(row=row+i, column=1).font = header_font
                ws.cell(row=row+i, column=2).fill = header_fill
                ws.cell(row=row+i, column=2).font = header_font
        
        # Status Distribution
        row += len(metrics) + 2
        ws.cell(row=row, column=1, value="Status Distribution")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        
        row += 1
        status_dist = summary.get('status_distribution', {})
        for status, count in status_dist.items():
            ws.cell(row=row, column=1, value=status)
            ws.cell(row=row, column=2, value=count)
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    @classmethod
    def _create_providers_sheet(cls, ws, results: Dict):
        """Create providers results worksheet"""
        from openpyxl.styles import Font, PatternFill
        
        headers = ['Provider ID', 'Name', 'Specialty', 'Status', 'Quality Score', 
                   'Priority', 'Issues Count', 'Phone Valid', 'Address Valid', 
                   'NPI Valid', 'License Valid']
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Add data rows
        provider_results = results.get('results', [])
        for row_idx, result in enumerate(provider_results, 2):
            provider = result.get('provider', {})
            report = result.get('report', {})
            validation = result.get('validation', {})
            validation_summary = report.get('validation_summary', {})
            
            ws.cell(row=row_idx, column=1, value=provider.get('provider_id', ''))
            ws.cell(row=row_idx, column=2, value=f"{provider.get('first_name', '')} {provider.get('last_name', '')}")
            ws.cell(row=row_idx, column=3, value=provider.get('specialty', ''))
            ws.cell(row=row_idx, column=4, value=report.get('overall_status', ''))
            ws.cell(row=row_idx, column=5, value=provider.get('data_quality_score', 0))
            ws.cell(row=row_idx, column=6, value=report.get('priority', ''))
            ws.cell(row=row_idx, column=7, value=len(provider.get('issues_found', [])))
            ws.cell(row=row_idx, column=8, value=validation_summary.get('phone', ''))
            ws.cell(row=row_idx, column=9, value=validation_summary.get('address', ''))
            ws.cell(row=row_idx, column=10, value=validation_summary.get('npi', ''))
            ws.cell(row=row_idx, column=11, value=validation_summary.get('license', ''))
        
        # Adjust column widths
        for col in range(1, 12):
            ws.column_dimensions[chr(64+col)].width = 15
    
    @classmethod
    def _create_issues_sheet(cls, ws, results: Dict):
        """Create issues summary worksheet"""
        from openpyxl.styles import Font, PatternFill
        
        headers = ['Provider ID', 'Provider Name', 'Issue', 'Severity', 'Field', 'Recommended Action']
        
        header_fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        row_idx = 2
        for result in results.get('results', []):
            provider = result.get('provider', {})
            report = result.get('report', {})
            quality = result.get('quality', {})
            
            provider_id = provider.get('provider_id', '')
            provider_name = f"{provider.get('first_name', '')} {provider.get('last_name', '')}"
            
            # Add red flags
            for flag in quality.get('red_flags', []):
                ws.cell(row=row_idx, column=1, value=provider_id)
                ws.cell(row=row_idx, column=2, value=provider_name)
                ws.cell(row=row_idx, column=3, value=flag.get('type', ''))
                ws.cell(row=row_idx, column=4, value=flag.get('severity', ''))
                ws.cell(row=row_idx, column=5, value=flag.get('field', ''))
                ws.cell(row=row_idx, column=6, value=flag.get('description', ''))
                row_idx += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 40
    
    @classmethod
    def to_csv(cls, results: Dict[str, Any]) -> str:
        """Export validation results to CSV format"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            'Provider ID', 'First Name', 'Last Name', 'Specialty', 'NPI',
            'Phone', 'Address', 'City', 'State', 'Zip',
            'Status', 'Quality Score', 'Priority', 'Issues Count',
            'Phone Valid', 'Address Valid', 'NPI Valid', 'License Valid',
            'Issues'
        ])
        
        # Data rows
        for result in results.get('results', []):
            provider = result.get('provider', {})
            report = result.get('report', {})
            validation_summary = report.get('validation_summary', {})
            
            writer.writerow([
                provider.get('provider_id', ''),
                provider.get('first_name', ''),
                provider.get('last_name', ''),
                provider.get('specialty', ''),
                provider.get('npi', ''),
                provider.get('phone', ''),
                provider.get('address', ''),
                provider.get('city', ''),
                provider.get('state', ''),
                provider.get('zip_code', ''),
                report.get('overall_status', ''),
                provider.get('data_quality_score', 0),
                report.get('priority', ''),
                len(provider.get('issues_found', [])),
                validation_summary.get('phone', ''),
                validation_summary.get('address', ''),
                validation_summary.get('npi', ''),
                validation_summary.get('license', ''),
                '; '.join(provider.get('issues_found', []))
            ])
        
        return output.getvalue()
    
    @classmethod
    def to_json(cls, results: Dict[str, Any], pretty: bool = True) -> str:
        """Export validation results to JSON format"""
        if pretty:
            return json.dumps(results, indent=2, default=str)
        return json.dumps(results, default=str)
